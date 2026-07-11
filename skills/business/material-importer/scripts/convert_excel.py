#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1.0"]
# ///
# ─── How to run ───
# 1. Install uv: curl -LsSf https://astral.sh/uv/install.sh | sh
# 2. Run: uv run convert_excel.py <input_dir> <output_dir>
# 3. Example: uv run convert_excel.py incoming/海鼎模板 raw/海鼎模板
# ──────────────────
"""Excel 工作簿 → 按复杂度分流为 Markdown(简单) 或 CSV(复杂)。

核心原则：一个源文件只产出一种格式，不混 .md 和 .csv。

路由规则（文件级）：
  - 该文件所有 sheet 都 ≤8 列 且 无合并 且 仅 1 个 sheet → Markdown table
  - 其余（任一 sheet 超阈值 / 有合并 / 多 sheet）→ CSV（每 sheet 一个）

技术处理：
  - 合并单元格值传播（openpyxl 只有左上角有值）
  - 双行表头展平（R1 分组合并 + R2 字段名 → 单行）
  - Excel 序列号日期 → ISO 字符串
  - .xls / 误标 .xlsx（OLE2）自动用 libreoffice 转换
"""

from __future__ import annotations

import csv
import os
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

SIMPLE_MAX_COLS = 8
EXCEL_EPOCH = datetime(1899, 12, 30)
TEMP_CONVERT = Path("/tmp/opencode/excel-convert")


def is_ole2(path: Path) -> bool:
    with open(path, "rb") as f:
        return f.read(8) == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def ensure_xlsx(path: Path) -> Path:
    """.xls 或误标 .xlsx(OLE2) → 用 libreoffice 转为真 .xlsx 到临时目录。"""
    if path.suffix == ".xlsx" and not is_ole2(path):
        return path

    TEMP_CONVERT.mkdir(parents=True, exist_ok=True)
    stem = path.stem
    target = TEMP_CONVERT / f"{stem}.xlsx"
    if not target.exists():
        subprocess.run(
            ["soffice", "--headless", "--convert-to", "xlsx",
             "--outdir", str(TEMP_CONVERT), str(path)],
            capture_output=True, timeout=60,
        )
    return target if target.exists() else path


def excel_serial_to_date(val: object) -> str:
    if isinstance(val, int | float) and 1 < val < 100000:
        try:
            return (EXCEL_EPOCH + timedelta(days=val)).strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            pass
    return str(val)


def normalize_cell(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = excel_serial_to_date(val)
    return s.strip() if isinstance(s, str) else str(s)


def read_sheet_data(ws, max_rows: int = 20000):
    """读取 sheet，传播合并值，展平双行表头。返回 (headers, rows)。"""
    merged_ranges = list(ws.merged_cells.ranges)

    merge_values: dict[tuple[int, int], object] = {}
    for mr in merged_ranges:
        top_left = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_values[(r, c)] = top_left

    all_rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=False):
        values = []
        for cell in row:
            pos = (cell.row, cell.column)
            raw = merge_values.get(pos, cell.value)
            values.append(normalize_cell(raw))
        if any(v for v in values):
            all_rows.append(values)

    if not all_rows:
        return [], []

    max_len = max(len(r) for r in all_rows)
    for r in all_rows:
        while len(r) < max_len:
            r.append("")

    # 双行表头展平
    if len(all_rows) >= 2 and len(merged_ranges) > 0:
        row1_has_col_merge = any(
            mr.min_row == 1 and mr.max_row == 1 and mr.max_col > mr.min_col
            for mr in merged_ranges
        )
        if row1_has_col_merge:
            headers: list[str] = []
            for h1, h2 in zip(all_rows[0], all_rows[1], strict=False):
                headers.append(h2 if h2 else (h1 if h1 else ""))
            return headers, all_rows[2:]

    return all_rows[0], all_rows[1:]


def sheet_is_simple(headers: list[str], has_merge: bool) -> bool:
    non_empty = sum(1 for h in headers if h)
    return non_empty <= SIMPLE_MAX_COLS and not has_merge


def file_is_simple(xlsx_path: Path) -> bool:
    """文件级判定：所有 sheet 都简单 且 仅 1 个 visible sheet → True。"""
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    visible_sheets = [
        ws for ws in wb.worksheets
        if getattr(ws, "state", "visible") == "visible"
    ]
    if len(visible_sheets) != 1:
        wb.close()
        return False

    headers, _ = read_sheet_data(visible_sheets[0])
    has_merge = len(list(visible_sheets[0].merged_cells.ranges)) > 0
    wb.close()
    return sheet_is_simple(headers, has_merge)


def write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def write_md(path: Path, title: str, headers: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [f"# {title}\n"]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for row in rows:
        lines.append("| " + " | ".join(row) + " |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def process_file(src: Path, output_dir: Path) -> str:
    """处理一个文件，返回产出的格式（'Markdown' / 'CSV' / 'skip'）。"""
    if src.suffix == ".docx":
        return "skip"

    xlsx = ensure_xlsx(src)
    if is_ole2(xlsx):
        return "skip"

    stem = src.stem
    simple = file_is_simple(xlsx)

    wb = load_workbook(xlsx, data_only=True, read_only=False)
    visible = [
        ws for ws in wb.worksheets
        if getattr(ws, "state", "visible") == "visible"
    ]

    if simple and len(visible) == 1:
        headers, rows = read_sheet_data(visible[0])
        if not headers:
            wb.close()
            return "skip"
        write_md(output_dir / f"{stem}.md", stem, headers, rows)
        wb.close()
        return f"Markdown({sum(1 for h in headers if h)}c/{len(rows)}r)"

    # 复杂：每 sheet 一个 CSV
    csv_dir = output_dir / "csv"
    for ws in visible:
        headers, rows = read_sheet_data(ws)
        if not headers:
            continue
        safe = ws.title.replace("/", "_").replace(" ", "_")
        name = f"{stem}_{safe}.csv" if len(visible) > 1 else f"{stem}.csv"
        write_csv(csv_dir / name, headers, rows)
    wb.close()
    return f"CSV({len(visible)}s)"


def main() -> int:
    if len(sys.argv) < 3:
        print(f"用法: {sys.argv[0]} <input_dir> <output_dir>")
        return 1

    input_dir = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    output_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(
        f for f in os.listdir(input_dir)
        if f.endswith((".xlsx", ".xls"))
    )
    print(f"处理 {len(files)} 个 Excel 文件...\n")

    for filename in files:
        result = process_file(input_dir / filename, output_dir)
        symbol = "✓" if result != "skip" else "⊘"
        print(f"  {symbol} {filename}: {result}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
