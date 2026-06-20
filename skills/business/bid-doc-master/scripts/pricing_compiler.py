"""报价单编译器 — YAML 配置 → 标准 .xlsx 报价单

用法:
    cd skills/business/bid-doc-master
    uv run scripts/pricing_compiler.py <报价配置.yaml> [--output 输出路径.xlsx]

YAML 格式见 materials/references/ 下的报价模板。
"""

import argparse
import sys
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_pricing(config: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRM报价"

    client = config.get("client", "")
    project = config.get("project", "")
    deploy_mode = config.get("mode", "saas")
    tax_rate = config.get("tax_rate", 0.06)
    bidder = config.get("bidder", "广州市蓝联科技有限公司")

    mode_label = "SAAS模式" if deploy_mode == "saas" else "私有化部署"

    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="微软雅黑", size=14, bold=True)
    sub_font = Font(name="微软雅黑", size=10)
    summary_font = Font(name="微软雅黑", size=10, bold=True)
    summary_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    remark_font = Font(name="微软雅黑", size=9, color="666666")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:J1")
    ws["A1"] = f"蓝联科技-智慧商圈会员营销CRM系统-{mode_label}报价单"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"服务商：{bidder}"
    ws["A2"].font = sub_font

    headers = [
        "序号", "名称", "内容说明",
        "首项目单价(元)", "数量", "首项目报价(元)",
        "新增项目单价(元)", "数量", "新增项目报价(元)",
        "备注",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    current_row = 4
    categories = config.get("categories", [])
    if not categories:
        categories = [{"name": "全部项目", "items": config.get("items", [])}]

    first_total = 0
    new_total = 0

    for cat in categories:
        cat_name = cat.get("name", "")
        items = cat.get("items", [])

        if cat_name:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            cell = ws.cell(row=current_row, column=1, value=cat_name)
            cell.font = Font(name="微软雅黑", size=10, bold=True)
            cell.fill = PatternFill(start_color="E8EDF3", end_color="E8EDF3", fill_type="solid")
            cell.alignment = left_wrap
            cell.border = thin_border
            current_row += 1

        for item in items:
            item_id = item.get("id", "")
            name = item.get("name", "")
            desc = item.get("description", "")
            first_unit = item.get("first_unit_price", 0) or 0
            first_qty = item.get("first_qty", 1) if first_unit else 0
            first_price = first_unit * first_qty if first_unit else 0
            new_unit = item.get("new_unit_price", 0) or 0
            new_qty = item.get("new_qty", 1) if new_unit else 0
            new_price = new_unit * new_qty if new_unit else 0
            remark = item.get("remark", "")
            optional = item.get("optional", False)

            if optional:
                remark = f"[可选] {remark}" if remark else "[可选]"

            first_total += first_price
            new_total += new_price

            row_data = [
                item_id, name, desc,
                first_unit if first_unit else "", first_qty if first_unit else "",
                first_price if first_price else "",
                new_unit if new_unit else "", new_qty if new_unit else "",
                new_price if new_price else "",
                remark,
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = sub_font
                cell.border = thin_border
                if col in (1, 4, 5, 6, 7, 8, 9, 10):
                    cell.alignment = center
                elif col in (2, 3):
                    cell.alignment = left_wrap
                if col in (4, 6, 7, 9) and isinstance(val, (int, float)) and val:
                    cell.number_format = '#,##0'
            current_row += 1

    tax_note = f"含税({int(tax_rate * 100)}%)"
    next_year_items = config.get("next_year", {})
    next_year_total = next_year_items.get("total", 0)

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.cell(row=current_row, column=1, value="首年费用合计").font = summary_font
    ws.cell(row=current_row, column=1).fill = summary_fill
    ws.cell(row=current_row, column=1).alignment = center
    ws.cell(row=current_row, column=1).border = thin_border

    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    cell = ws.cell(row=current_row, column=4, value=first_total)
    cell.font = summary_font
    cell.fill = summary_fill
    cell.alignment = right_align
    cell.border = thin_border
    cell.number_format = '#,##0'

    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=9)
    cell = ws.cell(row=current_row, column=7, value=new_total)
    cell.font = summary_font
    cell.fill = summary_fill
    cell.alignment = right_align
    cell.border = thin_border
    cell.number_format = '#,##0'

    ws.cell(row=current_row, column=10, value=tax_note).font = remark_font
    ws.cell(row=current_row, column=10).fill = summary_fill
    ws.cell(row=current_row, column=10).border = thin_border
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.cell(row=current_row, column=1, value="次年费用合计").font = summary_font
    ws.cell(row=current_row, column=1).fill = summary_fill
    ws.cell(row=current_row, column=1).alignment = center
    ws.cell(row=current_row, column=1).border = thin_border

    next_val = next_year_total or (new_total // 3 if new_total else 0)
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    cell = ws.cell(row=current_row, column=4, value=next_val)
    cell.font = summary_font
    cell.fill = summary_fill
    cell.alignment = right_align
    cell.border = thin_border
    cell.number_format = '#,##0'

    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=10)
    ws.cell(row=current_row, column=7, value=next_year_items.get("note", "年度授权+售后+云服务")).font = remark_font
    ws.cell(row=current_row, column=7).fill = summary_fill
    ws.cell(row=current_row, column=7).border = thin_border
    current_row += 2

    service_notes = config.get("service_notes", [])
    if service_notes:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=1, value="服务说明：").font = Font(name="微软雅黑", size=9, bold=True)
        current_row += 1
        for note in service_notes:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1, value=note).font = remark_font
            current_row += 1

    col_widths = [8, 18, 40, 14, 6, 14, 14, 6, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 25

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="报价单编译器：YAML → .xlsx")
    parser.add_argument("config", help="报价配置 YAML 文件路径")
    parser.add_argument("--output", "-o", help="输出 .xlsx 路径")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    with open(args.config, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    output = args.output
    if not output:
        stem = Path(args.config).stem
        deploy = config.get("mode", "saas")
        client = config.get("client", "客户")
        output = f"报价单_{deploy}_{client}.xlsx"

    generate_pricing(config, output)

    if args.verbose:
        items_count = sum(len(c.get("items", [])) for c in config.get("categories", [{"items": config.get("items", [])}]))
        print(f"✅ 报价单生成成功!")
        print(f"📁 {output}")
        print(f"📊 项目数: {items_count}")
        print(f"🏢 客户: {config.get('client', '?')}")
        print(f"📦 模式: {config.get('mode', 'saas')}")


if __name__ == "__main__":
    main()
