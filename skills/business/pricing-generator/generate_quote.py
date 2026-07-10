"""Generate Excel quote (.xlsx) for pricing-generator skill.

参考正祥报价单格式（深蓝表头 / 浅红汇总 / 微软雅黑 / 列宽规范），
生成双 sheet Excel：Sheet1=报价单，Sheet2=功能清单。

Usage:
    uv run python generate_quote.py --customer XX资产公司 --product MI --mode SAAS [--date 20260705]

环境变量：
    LANLNK_BASE  素材库根目录（默认 /opt/code/docs/lanlnk）
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# A4 纸张代码（Excel paperSize）
PAPER_A4 = 9

# === 样式常量（参考正祥报价单）===
FONT_NAME = "微软雅黑"
COLOR_HEADER_BG = "1E5A8A"          # 深蓝表头填充
COLOR_HEADER_FG = "FFFFFF"          # 表头白字
COLOR_SECTION_BG = "E4F0FA"         # 浅蓝小节标题
COLOR_HIGHLIGHT_BG = "F9E5DD"       # 浅红汇总/合计行
COLOR_AMOUNT_FG = "D46B43"          # 金额高亮橙
COLOR_INFO_FG = "555555"            # 信息行灰字
COLOR_FOOTER_FG = "888888"          # 页脚浅灰

_THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 报价单 Sheet 列宽（6 列）
QUOTE_COL_WIDTHS = {"A": 8, "B": 22, "C": 50, "D": 14, "E": 14, "F": 18}
# 功能清单 Sheet 列宽（3 列）
MODULE_COL_WIDTHS = {"A": 6, "B": 30, "C": 80}

# === 服务商信息 ===
PROVIDER = {
    "name": "广州市蓝联科技有限公司",
    "contact": "顾为国 / 13711775158",
}

# === MI 商管系统数据（v1 硬编码，用于验证；后续版本改为读模板动态生成）===
MI_DATA: dict[str, Any] = {
    "product_name": "MI 商管系统",
    "product_label": "MI",
    # 标准产品（必选）：序号 / 名称 / 内容说明 / 首年报价 / 次年报价 / 备注
    "standard_items": [
        ("1.1", "商管平台 SAAS 租用",
         "含 9 大模块标准功能（招商/资源/合同/计费/运营/物业/报表/移动端/系统管理），"
         "为写字楼+商业项目搭建招商经营闭环",
         20000, 20000, "年租用费"),
        ("1.2", "实施服务",
         "项目启动（业务调研/需求确认/计划制定）、方案设计（部署/蓝图/数据准备）、"
         "上线运行（切换/检查/问题跟踪）、项目验收（文档/售后交接）",
         30000, 0, "首年一次性，含 15 人天"),
        ("1.3", "年度售后服务",
         "问题排查/修复/安全漏洞修补/技术维护",
         0, 0, "首年赠送，含在租用费"),
    ],
    # 定制开发（可选）：序号 / 名称 / 内容说明 / 首年报价 / 次年报价 / 备注
    "optional_items": [
        ("2.1", "定制开发-线索确权规则",
         "公海池/线索保护期/首访确权判定/多渠道撞客处理/佣金结算确权规则",
         22000, 0, "11 人天 × 2000/天；P0 首期，可选"),
        ("2.2", "定制开发-写字楼房屋售卖",
         "产权交易全流程：售卖合同/售价管理/认购记录/房款分期/业财打通",
         60000, 0, "30 人天 × 2000/天；P1 二期，可选"),
    ],
    # 第三方集成（不报价）：集成项 / 说明 / 计费方式
    "third_party": [
        ("在线支付网关", "微信/支付宝/对公在线缴费", "客户指定支付网关后单独计费"),
        ("第三方电子签", "法大大/e签宝等电子合同签署", "按第三方报价 + 集成人天"),
        ("AI 客流系统对接", "客流摄像头数据接入（硬件由客流系统供应商提供）",
         "按客流系统报价 + 接口对接人天"),
        ("活动效果自动复盘/投放归因", "活动复盘报告/投放 ROI 归因分析",
         "由蓝联 langchat AI Skills 承接，另报价"),
    ],
    # 方案对比：序号 / 方案 / 首年汇总 / 首年优惠价(留空) / 次年汇总 / 次年优惠价(留空) / 适用
    "plans": [
        ("5.1", "方案 A：标准产品（无二开，纯标准 SAAS）",
         50000, "", 20000, "",
         "客户首期只要标准招商经营闭环，确权规则用标准配置，无售卖业务"),
        ("5.2", "方案 B：标准 + 确权规则（P0）",
         69800, "", 20000, "",
         "客户有多渠道招商（代理/中介/经纪人），需要标准化判客确权和佣金结算规则"),
        ("5.3", "方案 C：标准 + 确权 + 售卖（P0+P1）",
         123800, "", 20000, "",
         "客户写字楼既有租赁又有产权售卖，需要完整租售管理"),
    ],
    # 功能模块（Sheet2）：模块名 / 功能描述
    "modules": [
        ("资源管理",
         "铺位管理（资产性流程审批/非资产性直接保存）、空间结构管理（楼宇/楼层/位置/区域）、"
         "铺位审批、拆分合并历史、多经点位管理、车位管理、可租面积管理、"
         "租控可视化、空置地图、项目资源"),
        ("招商管理",
         "品牌库（分级管理/黑名单/集团共享）、品牌入库审批（项目→城市→总部）、"
         "招商计划（年度/月度）、招商任务、招商绩效、招商预警、意向商户/意向单元管理、"
         "渠道管理、客户管理（客户跟进/客户转介）、年度指标"),
        ("合同管理",
         "租约管理（创建/详情/列表）、合同生成（模板选择/文本编辑/附件/打印预览/审批→台账）、"
         "合同模板管理、租赁条件（固定租金/提成租金/物业管理费/宣传推广费/保证金/最低营业额/"
         "滞纳金规则/含税不含税/协议类型/统收非统收）、合同变更（主体/位置/期限/费用）、"
         "合同终止（租约解除/应付未付）、合同清算（保证金/费用联动）、租金决策对比、提成租金预览"),
        ("财务管理",
         "计费出账、应收管理/应收审核、收费方案管理/收费审核、账单生成与多渠道通知、"
         "收款核销、银行对账、退款/减免审批、押金余额、发票（申请/审核/明细）、"
         "税务导出/税率配置、财务配置/财务期间管理、未知付款/资金勾兑"),
        ("运营管理",
         "销售管理/销售监控/销售复核/销售调整、商户台账/商户档案/商户注册、"
         "装修申请管理、商户履约评价、巡检任务/巡检详情、推广管理/推广合同/推广台账、估算填报"),
        ("物业服务",
         "报修工单管理、工单/工单详情、服务申请（提交/详情）、"
         "工程条件管理、设备管理、智能设备管理、值班日志"),
        ("报表与看板",
         "仪表盘、工作台/工作台页面/首页、项目概览、报表管理器、"
         "自定义报表设计器、Excel 导入导出、综合报表、"
         "客流报表、商铺可视化分析、打印输出/打印模板、凭证预览、收据管理"),
        ("移动端",
         "管理端：工作台、审批/审批详情、公告（管理/通知/统计/详情）、巡检任务、值班日志、"
         "消息中心、抄表扫码、项目概览。\n"
         "租户端：注册/登录、我的、账单/账单详情/付款记录、服务申请/提交服务申请/申请详情、"
         "销售填报/日销售/填报记录、活动资讯/活动详情、发票申请、违约单/违约单详情、"
         "消息中心/消息详情/通知中心/通知详情"),
        ("系统管理",
         "组织架构、用户管理、菜单管理、权限控制、"
         "数据隔离、工作流管理（工作流审批/工作流可视化编辑/"
         "工作流通知/工作流定义）、年度指标、主数据管理、通知记录"),
    ],
    # 服务说明脚注
    "service_notes": [
        "1. 服务承诺：提供首次上线的后台数据切换，包括铺位/合同/应收数据初始化、"
        "历史数据迁移、系统配置；",
        "2. 二开单价：未来新需求，二开人天单价按 2,000 元/人天结算；",
        "3. SAAS 服务范围：含平台运维、安全更新、版本升级；不含定制开发和数据导出定制；",
        "4. 实施人天：首年实施含 15 人天（项目启动+方案设计+上线+验收），"
        "超出部分按 2,000 元/人天另计；",
        "5. 付款方式：首年签订合同时支付 50%，验收后支付 50%；次年租用费按年续费支付。",
    ],
    "standard_first_year_total": 50000,
    "standard_next_year_total": 20000,
}


CRM_DATA: dict[str, Any] = {
    "product_name": "CRM 会员营销系统",
    "product_label": "CRM",
    "standard_items": [
        ("1.1", "CRM 平台 SAAS 租用",
         "智慧商圈会员营销 CRM 平台，含电子会员/会员卡券/会员积分/会员标签/会员等级/会员资产",
         60000, 30000, "年租用费"),
        ("1.2", "智慧商圈",
         "开通微信/支付宝智慧商圈，实现支付即积分/支付即会员/无感积分/停车积分",
         20000, 0, "必选，首年开通费"),
        ("1.3", "实施服务",
         "项目启动/方案设计/上线运行/项目验收",
         20000, 0, "首年一次性，含 10 人天"),
        ("1.4", "年度售后服务",
         "问题排查/修复/安全漏洞修补/技术维护",
         0, 0, "首年赠送，含在租用费"),
    ],
    "optional_items": [
        ("2.1", "停车对接",
         "CRM 对接停车系统，对接在线缴费/积分/停车券优惠停车等停车营销服务",
         10000, 0, "可选"),
        ("2.2", "美团对接",
         "CRM 对接美团系统，对接会员/积分/优惠券等营销服务",
         10000, 0, "可选"),
        ("2.3", "抖音对接",
         "CRM 对接抖音系统，对接会员/积分/优惠券等营销服务",
         10000, 0, "可选"),
    ],
    "third_party": [],
    "plans": [
        ("5.1", "方案 A：标准产品（CRM 平台 + 智慧商圈 + 实施 + 售后）",
         100000, "", 30000, "",
         "标准会员营销闭环"),
        ("5.2", "方案 B：标准 + 停车对接",
         110000, "", 30000, "",
         "含停车积分营销"),
        ("5.3", "方案 C：标准 + 全部对接（停车 + 美团 + 抖音）",
         130000, "", 30000, "",
         "含停车/美团/抖音平台打通"),
    ],
    "modules": [
        ("会员管理",
         "电子会员/会员卡券/会员积分管理/会员标签/会员等级/会员资产/会员档案"),
        ("智慧商圈",
         "微信/支付宝智慧商圈/支付即积分/支付即会员/无感积分/停车积分"),
        ("营销活动",
         "优惠券/满减/折扣/抽奖/拼团/秒杀/活动报名/活动审批/活动效果分析"),
        ("停车营销",
         "停车券/积分抵扣/在线缴费/停车月卡/停车积分规则"),
        ("数据分析",
         "会员画像/消费分析/积分分析/活跃度分析/流失预警/会员价值分级"),
        ("消息中心",
         "短信/模板消息/订阅消息/推送通知/会员触达/消息模板管理"),
        ("移动端",
         "会员小程序/会员中心/电子会员卡/卡券包/积分商城/在线缴费"),
        ("系统管理",
         "组织架构/权限管理/门店管理/数据隔离/日志审计/系统配置"),
    ],
    "modules_xlsx": "/opt/code/docs/lanlnk/out/proposals/正祥会员系统/蓝联CRM功能清单.xlsx",
    "service_notes": [
        "1. 服务承诺：提供首次上线的后台数据切换，包括会员数据切入、积分数据切入、订单数据切入、历史会员资产备份；",
        "2. 二开单价：未来新需求，二开人天单价按 2,000 元/人天结算；",
        "3. SAAS 服务范围：含平台运维、安全更新、版本升级；不含定制开发和数据导出定制；",
        "4. 实施人天：首年实施含 10 人天（项目启动+方案设计+上线+验收），超出部分按 2,000 元/人天另计；",
        "5. 付款方式：首年签订合同时支付 50%，验收后支付 50%；次年租用费按年续费支付。",
    ],
    "standard_first_year_total": 100000,
    "standard_next_year_total": 30000,
}


# === AI 岗位 Skill 数据（动态生成，按 positions 岗位数计算费用）===
AI_POSITION_SKILLS: list[tuple[str, str]] = [
    ("营运分析 Skill",
     "经营日报/周报自动生成、异常预警、风险商户清单；P0 主线岗位"),
    ("客服分流 Skill",
     "高频 FAQ 应答、会员查询、转人工工单；P0 主线岗位"),
    ("招商研究 Skill",
     "品牌初筛画像、铺位匹配、招商漏斗周报；P0 主线岗位"),
    ("财务账龄预警 Skill",
     "账龄分析、催缴优先级、催缴跟踪；P0 备选岗位"),
    ("工程工单协同 Skill",
     "工单派单、SLA 预警、超时升级；P0 备选岗位"),
    ("企划岗 Skill 组合",
     "活动复盘、会员运营、内容生产、投放归因；P1 扩展岗位"),
]

AI_SERVICE_NOTES: list[str] = [
    "1. 2 岗位起卖：单岗位难以形成经营闭环，至少 2 个岗位才能体现 AI 增强价值；",
    "2. 首年含实施：每岗位首年 1 万含平台部署 + 知识库建设 + 工作流配置 + 试点运行"
    "（约 5 人天/岗位）；",
    "3. 次年运营服务费固定 2 万：不论购买几个岗位，次年运营服务费均为 2 万。"
    "AI 岗位 Skill 卖的是运营服务——周复盘/命中率调优/知识库更新/效果看板/SLA 保障/持续优化，"
    "确保 Skill 越用越准；",
    "4. 新岗位 Skill 定制：如需当前 6 个以外的岗位 Skill，按 2,000 元/人天定制开发；",
    "5. 数据接入：客户需提供商管系统/CRM/客流等数据源接口；"
    "LnkAgent 本地部署确保数据安全；",
    "6. 退出机制：试点 90 天后效果不达预期可停止，不绑定长期合同。",
]


def build_ai_data(positions: int) -> dict[str, Any]:
    """根据岗位数构建 AI Skills 报价数据。

    定价规则：
    - 2 岗起卖（positions < 2 强制为 2）
    - 首年 = positions × 10,000
    - 次年 = 20,000（固定运营服务费）
    """
    positions = max(2, min(positions, 6))  # 2岗起卖，当前最多6个岗位
    first_year = positions * 10000
    next_year = 20000

    selected = AI_POSITION_SKILLS[:positions]
    selected_names = "、".join(
        s[0].replace(" Skill", "").replace(" 组合", "") for s in selected
    )

    return {
        "product_name": "LnkAgent AI岗位Skill增强服务",
        "product_label": "AI",
        "mode_desc": "报价模式：SAAS 运营服务（首年含岗位 Skill 配置 + 上线试运营；次年固定运营服务费 2 万）",
        "standard_items": [
            ("1.1", "LnkAgent 平台 SAAS",
             "AI Skills 编排平台（OrchestratorAgent + langchat + LnkChatBI 三引擎），"
             "本地部署或 SAAS 托管，确保数据安全",
             0, 0, "含在岗位费"),
            ("1.2", f"岗位 Skill × {positions}",
             f"已选 {positions} 个岗位：{selected_names}。"
             "每个 Skill 含场景配置 + 知识库建设 + 工作流上线 + 试点运行",
             first_year, 0, f"{positions} × 10,000；2 岗起卖"),
            ("1.3", "实施服务",
             "数据接入 + 知识库建设 + 工作流配置 + 试点辅导（约 5 人天/岗位）",
             0, 0, "含在岗位费"),
            ("1.4", "年度运营服务",
             "周复盘 + 命中率调优 + 知识库更新 + 效果看板 + SLA 保障 + 持续优化",
             0, next_year, "次年固定 2 万；首年含试运营"),
        ],
        "optional_items": [
            ("2.1", "新岗位 Skill 定制开发",
             "当前 6 个岗位 Skill 以外的定制需求（如物业岗/招商谈判岗等），按人天结算",
             0, 0, "2,000 元/人天；按需"),
        ],
        "third_party": [
            ("数据源接入", "商管系统/CRM/客流等数据接口",
             "客户提供接口，LnkAgent 本地部署对接"),
            ("大模型 API", "GLM/DeepSeek/通义等大模型推理 API",
             "客户自购或蓝联代采，按 token 计费"),
        ],
        "plans": [
            ("5.1", "入门 2岗（营运分析 + 客服分流）",
             20000, "", 20000, "",
             "最小闭环：经营日报 + 客服减负"),
            ("5.2", "标准 3岗 ⭐（营运 + 客服 + 招商研究）",
             30000, "", 20000, "",
             "推荐起步：P0 三主 Skill"),
            ("5.3", "增强 4岗（三主 + 财务账龄/工程工单）",
             40000, "", 20000, "",
             "P0 三主 + 一备选"),
            ("5.4", "完整 6岗（全部 6 个岗位 Skill）",
             60000, "", 20000, "",
             "全岗位 AI 增强"),
        ],
        "modules": AI_POSITION_SKILLS,
        "service_notes": AI_SERVICE_NOTES,
        "standard_first_year_total": first_year,
        "standard_next_year_total": next_year,
    }


def merge_product_data(products: list[dict[str, Any]]) -> dict[str, Any]:
    """合并多个产品的报价数据。

    合并规则：
    - standard_items: 全部合并，序号重编 1.1/1.2/...，名称加【产品标签】前缀
    - optional_items: 按名称去重，序号重编 2.1/2.2/...
    - third_party: 按名称去重
    - plans: 动态生成3个方案（标准组合 / 标准组合+P0二开 / 标准组合+全部二开）
    - modules: 全部合并，名称加【产品标签】前缀
    - service_notes: 按内容去重
    - totals: 首年/次年分别求和
    """
    merged_label = "+".join(p["product_label"] for p in products)
    merged_name = "+".join(p["product_name"] for p in products)

    standard_items: list[tuple] = []
    seq = 1
    for p in products:
        for item in p["standard_items"]:
            standard_items.append((
                f"1.{seq}", f"【{p['product_label']}】{item[1]}",
                item[2], item[3], item[4], item[5],
            ))
            seq += 1

    optional_items: list[tuple] = []
    seen_opt: set[str] = set()
    seq = 1
    for p in products:
        for item in p["optional_items"]:
            if item[1] in seen_opt:
                continue
            seen_opt.add(item[1])
            optional_items.append((
                f"2.{seq}", f"【{p['product_label']}】{item[1]}",
                item[2], item[3], item[4], item[5],
            ))
            seq += 1

    third_party: list[tuple] = []
    seen_tp: set[str] = set()
    for p in products:
        for item in p["third_party"]:
            if item[0] in seen_tp:
                continue
            seen_tp.add(item[0])
            third_party.append(item)

    modules: list[tuple] = []
    for p in products:
        for mod in p["modules"]:
            modules.append((f"【{p['product_label']}】{mod[0]}", mod[1]))

    service_notes: list[str] = []
    seen_sn: set[str] = set()
    for p in products:
        for note in p["service_notes"]:
            if note in seen_sn:
                continue
            seen_sn.add(note)
            service_notes.append(note)

    std_first = sum(p["standard_first_year_total"] for p in products)
    std_next = sum(p["standard_next_year_total"] for p in products)

    p0_first = sum(
        it[3] for it in optional_items if "P0" in it[5] or "必需" in it[5]
    )
    all_opt_first = sum(it[3] for it in optional_items)

    plans = [
        ("5.1", "方案 A：标准产品组合（无二开）",
         std_first, "", std_next, "",
         "所有产品标准组合起步，无任何二开"),
        ("5.2", "方案 B：标准组合 + P0 必需二开",
         std_first + p0_first, "", std_next, "",
         "在标准组合基础上加 P0 必需二开"),
        ("5.3", "方案 C：标准组合 + 全部二开",
         std_first + all_opt_first, "", std_next, "",
         "在标准组合基础上加全部可选二开"),
    ]

    return {
        "product_name": merged_name,
        "product_label": merged_label,
        "standard_items": standard_items,
        "optional_items": optional_items,
        "third_party": third_party,
        "plans": plans,
        "modules": modules,
        "service_notes": service_notes,
        "standard_first_year_total": std_first,
        "standard_next_year_total": std_next,
    }


def get_lanlnk_base() -> Path:
    return Path(os.environ.get("LANLNK_BASE", "/opt/code/docs/lanlnk"))


def parse_devkit(path: Path) -> list[tuple[str, str, str, int, int, str]]:
    """解析 requirement-evaluator 的需求评估报告，提取 MI 二开清单。

    返回 optional_items 格式：[(序号, 名称, 说明, 首年报价, 次年报价, 备注), ...]
    报价 = 人天 × 2000（pricing-generator 报价单价）
    """
    import re
    text = path.read_text(encoding="utf-8")
    items = []
    # 匹配二开清单表格行：| # | 二开项 | ... | 人天 | ...
    # 复杂度列可能带 markdown 加粗（**L**），用 \*{0,2} 兼容
    pattern = r"\|\s*(\d+)\s*\|\s*(.+?)\s*\|.*?\|\s*\*{0,2}([SML])\*{0,2}\s*\|\s*(\d+)\s*\|"
    for m in re.finditer(pattern, text):
        num, name, complexity, days = m.group(1), m.group(2).strip(), m.group(3), int(m.group(4))
        cost = days * 2000
        seq = f"2.{num}"
        items.append((
            seq, f"定制开发-{name}",
            f"{name}（{complexity} 级，{days} 人天）",
            cost, 0, f"{days} 人天 × 2,000/天；P{0 if complexity in ('S','M') else 1}"
        ))
    return items


# === 单元格样式辅助 ===
def style_cell(
    cell,
    *,
    size: int = 10,
    bold: bool = False,
    color: str | None = None,
    bg: str | None = None,
    h: str = "center",
    v: str = "center",
    wrap: bool = True,
    border: bool = True,
) -> None:
    cell.font = Font(
        name=FONT_NAME, size=size, bold=bold, color=color or "000000"
    )
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border:
        cell.border = BORDER_ALL


def write_title(ws, row: int, text: str, end_col: int = 6) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row, 1, text)
    style_cell(c, size=14, bold=True, h="center", v="center", border=False)
    ws.row_dimensions[row].height = 28


def write_info_row(ws, row: int, text: str, end_col: int = 6,
                   color: str = COLOR_INFO_FG) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row, 1, text)
    style_cell(c, size=9, color=color, h="left", v="center",
               border=False, wrap=False)
    ws.row_dimensions[row].height = 20


def write_section_title(ws, row: int, text: str, end_col: int = 6,
                        bg: str = COLOR_SECTION_BG) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row, 1, text)
    style_cell(c, size=11, bold=True, bg=bg, h="left", v="center", border=False)
    ws.row_dimensions[row].height = 22


def write_header_row(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        style_cell(c, size=10, bold=True,
                   color=COLOR_HEADER_FG, bg=COLOR_HEADER_BG)
    ws.row_dimensions[row].height = 22


def _estimate_height(values: list[Any], col_widths_chars: dict[int, int]) -> float:
    """根据内容显示宽度估算行高（中文/全角计 2、半角计 1）。"""
    max_lines = 1
    for i, v in enumerate(values, start=1):
        if not isinstance(v, str) or not v:
            continue
        width = col_widths_chars.get(i, 20)
        capacity = max(6.0, width * 0.95)
        for seg in str(v).split("\n"):
            seg_w = sum(2 if ord(ch) > 127 else 1 for ch in seg)
            lines = max(1, -(-seg_w // capacity))
            max_lines = max(max_lines, int(lines))
    return max(24.0, 17.0 * max_lines + 8.0)


def apply_a4_print(
    ws,
    *,
    orientation: str = "portrait",
    title_rows: str | None = None,
) -> None:
    """配置 A4 打印：宽度压 1 页、高度自动多页、水平居中、窄边距。"""
    ws.page_setup.paperSize = PAPER_A4
    ws.page_setup.orientation = orientation
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.3, footer=0.3
    )
    if title_rows:
        ws.print_title_rows = title_rows


def write_data_row(
    ws,
    row: int,
    values: list[Any],
    *,
    total: bool = False,
    bold: bool = False,
    highlight: bool = False,
) -> None:
    bg = COLOR_HIGHLIGHT_BG if highlight else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i, v)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            amt_color = COLOR_AMOUNT_FG if total else None
            style_cell(c, size=11 if total else 10,
                       bold=bold or total, color=amt_color,
                       bg=bg, h="center", v="center")
            c.number_format = "#,##0"
        else:
            h_align = "left" if i == 3 else "center"
            style_cell(c, size=11 if total else 10,
                       bold=bold or total, bg=bg, h=h_align, v="center")
    col_widths = {idx: QUOTE_COL_WIDTHS[col]
                  for idx, col in enumerate(["A", "B", "C", "D", "E", "F"], 1)}
    ws.row_dimensions[row].height = _estimate_height(values, col_widths)


# === Sheet 1：报价单 ===
def build_quote_sheet(
    wb: Workbook,
    data: dict[str, Any],
    customer: str,
    mode: str,
    date_str: str,
) -> None:
    ws = wb.active
    ws.title = f"{data['product_label']}报价"

    for col, w in QUOTE_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    end_col = 6
    r = 1

    # 标题
    title = f"蓝联科技-{data['product_name']}-{mode}报价单"
    write_title(ws, r, title, end_col); r += 1

    # 信息行
    info = (f"服务商：{PROVIDER['name']}     联系信息：{PROVIDER['contact']}     "
            f"客户：{customer}     报价日期：{date_str}     有效期：30 天")
    write_info_row(ws, r, info, end_col); r += 1

    # 模式说明
    if data.get("mode_desc"):
        mode_desc = data["mode_desc"]
    elif mode == "SAAS":
        mode_desc = "报价模式：SAAS（首年含软件租用费 + 实施费；次年只有年租用费）"
    else:
        mode_desc = "报价模式：私有化（首年含软件终生授权 + 实施费；次年售后可选）"
    write_info_row(ws, r, mode_desc, end_col); r += 1

    # 二开单价
    write_info_row(ws, r,
                   "二开单价：未来新需求，二开人天单价按 2,000 元/人天结算",
                   end_col); r += 1

    r += 1  # 空行

    # 一、标准产品报价（必选）
    write_section_title(ws, r, "一、标准产品报价（必选）", end_col); r += 1
    write_header_row(ws, r,
                     ["序号", "名称", "内容说明",
                      "首年报价(元)", "次年报价(元)", "备注"])
    r += 1
    for item in data["standard_items"]:
        write_data_row(ws, r, list(item)); r += 1
    # 标准汇总
    write_data_row(
        ws, r,
        ["", "标准产品合计", "",
         data["standard_first_year_total"],
         data["standard_next_year_total"],
         "首年/次年汇总"],
        total=True, highlight=True,
    ); r += 1

    r += 1

    # 二、第三方对接（可选，单独计费）
    write_section_title(ws, r, "二、第三方对接（可选，单独计费）", end_col); r += 1
    write_header_row(ws, r,
                     ["序号", "名称", "内容说明",
                      "首年报价(元)", "次年报价(元)", "备注"])
    r += 1
    for item in data["optional_items"]:
        write_data_row(ws, r, list(item)); r += 1

    r += 1

    # 三、第三方集成（另行计费，仅在 non-empty 时渲染）
    if data.get("third_party"):
        write_section_title(ws, r,
                            "三、第三方集成（另行计费，不在本报价内）",
                            end_col); r += 1
        write_header_row(ws, r,
                         ["序号", "集成项", "说明", "", "", "备注"]); r += 1
        for i, item in enumerate(data["third_party"], start=1):
            write_data_row(ws, r, [f"3.{i}", item[0], item[1], "", "", item[2]]); r += 1

        r += 1

    # 三、汇总区
    write_section_title(ws, r,
                        "三、汇总：报价合计",
                        end_col); r += 1
    write_header_row(ws, r,
                     ["序号", "项目", "首年汇总(元)", "首年优惠价",
                      "次年汇总(元)", "次年优惠价"]); r += 1
    write_data_row(ws, r,
                   ["3.1", "标准产品",
                    data["standard_first_year_total"], "",
                    data["standard_next_year_total"], ""],
                   highlight=True); r += 1
    opt_first = sum(it[3] for it in data["optional_items"])
    write_data_row(ws, r,
                   ["3.2", "+ 第三方对接（全选）", opt_first, "", 0, ""],
                   highlight=True); r += 1
    write_data_row(ws, r,
                   ["3.3", "合计",
                    data["standard_first_year_total"] + opt_first, "",
                    data["standard_next_year_total"], ""],
                   total=True, highlight=True); r += 1

    r += 1

    # 四、服务说明
    write_section_title(ws, r, "四、服务说明", end_col); r += 1
    for note in data["service_notes"]:
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=end_col)
        c = ws.cell(r, 1, note)
        style_cell(c, size=10, h="left", v="center", border=False)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1
    # 页脚
    ws.merge_cells(start_row=r, start_column=1,
                   end_row=r, end_column=end_col)
    c = ws.cell(r, 1,
                "服务说明：本报价仅供客户项目成员内部参阅，有效期 30 天，请勿外传。")
    style_cell(c, size=9, color=COLOR_FOOTER_FG, h="left", border=False)

    apply_a4_print(ws, orientation="portrait", title_rows="1:4")
    # 冻结前 4 行（标题/信息/模式/单价）
    ws.freeze_panes = "A5"


# === Sheet 2：功能清单 ===
def build_modules_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet(title=f"{data['product_label']}功能清单")

    xlsx_path = data.get("modules_xlsx")
    if xlsx_path and Path(xlsx_path).exists():
        _fill_modules_from_xlsx(ws, xlsx_path, data)
    else:
        _fill_modules_from_data(ws, data)


def _fill_modules_from_xlsx(ws, xlsx_path: str, data: dict[str, Any]) -> None:
    import openpyxl as _opx
    src_wb = _opx.load_workbook(xlsx_path, data_only=True)
    src_ws = src_wb.active

    col_widths = {"A": 4, "B": 10, "C": 13, "D": 18, "E": 50}
    col_width_map = {i: col_widths[c] for i, c in enumerate("ABCDE", 1)}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, f"{data['product_name']} 功能清单")
    style_cell(c, size=14, bold=True, h="center", v="center", border=False)
    ws.row_dimensions[r].height = 28
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(
        r, 1,
        "以下为标准功能清单（源自蓝联CRM功能清单.xlsx）。定价见「报价单」Sheet。",
    )
    style_cell(c, size=9, color=COLOR_INFO_FG, h="left", v="center", border=False)
    ws.row_dimensions[r].height = 20
    r += 2

    header_row: int | None = None
    data_started = False
    group_b_rows: list[int] = []
    group_c_rows: list[int] = []
    section_rows: list[int] = []
    last_data_row: int | None = None

    src_rows = list(src_ws.iter_rows(min_row=1, values_only=True))
    # 源文件第 1 行是其自带总标题，与上方生成标题重复，跳过
    for src_idx, row in enumerate(src_rows):
        if src_idx == 0:
            continue
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        values = list(row[:5]) + [None] * (5 - len(row[:5]))

        is_section_title = (
            values[0] is not None and str(values[0]).strip() != ""
            and all(v is None or str(v).strip() == "" for v in values[1:5])
        )

        if is_section_title:
            section_rows.append(r)
            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r, end_column=5)
            cell = ws.cell(r, 1, values[0])
            style_cell(cell, size=10, bold=True, h="left", v="center")
            ws.row_dimensions[r].height = 22
        else:
            is_header = (
                header_row is None
                and values[0] is not None
                and "序号" in str(values[0])
            )
            if is_header:
                header_row = r
            elif header_row is not None:
                data_started = True
            for col_idx in range(1, 6):
                v = values[col_idx - 1]
                cell = ws.cell(r, col_idx, v if v is not None else "")
                if col_idx == 1:
                    style_cell(cell, size=9, h="center", v="center")
                elif col_idx in (2, 3):
                    style_cell(cell, size=9, bold=True, h="left", v="center")
                else:
                    style_cell(cell, size=9, h="left", v="center")
            ws.row_dimensions[r].height = _estimate_height(
                [str(v) if v is not None else "" for v in values], col_width_map
            )
            if data_started:
                if values[1] not in (None, ""):
                    group_b_rows.append(r)
                if values[2] not in (None, ""):
                    group_c_rows.append(r)
                last_data_row = r
        r += 1

    def _merge_group(col_idx: int, anchor_rows: list[int]) -> None:
        for i, start in enumerate(anchor_rows):
            next_anchor = anchor_rows[i + 1] - 1 if i + 1 < len(anchor_rows) else last_data_row
            next_section = min(
                (s - 1 for s in section_rows if s > start), default=last_data_row
            )
            end = min(next_anchor, next_section)
            if end and end > start:
                ws.merge_cells(
                    start_row=start, start_column=col_idx,
                    end_row=end, end_column=col_idx,
                )

    _merge_group(2, group_b_rows)
    _merge_group(3, group_c_rows)

    top = header_row or 4
    apply_a4_print(ws, orientation="portrait", title_rows=f"1:{top}")
    ws.freeze_panes = f"A{top + 1}"


def _fill_modules_from_data(ws, data: dict[str, Any]) -> None:
    for col, w in MODULE_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1, f"{data['product_name']} 功能清单（按模块）")
    style_cell(c, size=14, bold=True, h="center", v="center", border=False)
    ws.row_dimensions[r].height = 28
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(
        r, 1,
        "以下为 SAAS 租用包含的标准功能模块。定价见「报价单」Sheet。",
    )
    style_cell(c, size=9, color=COLOR_INFO_FG, h="left", v="center", border=False)
    ws.row_dimensions[r].height = 20
    r += 2

    for i, h in enumerate(["序号", "模块名", "功能描述"], start=1):
        c = ws.cell(r, i, h)
        style_cell(c, size=10, bold=True,
                   color=COLOR_HEADER_FG, bg=COLOR_HEADER_BG)
    ws.row_dimensions[r].height = 22
    r += 1

    for i, (name, desc) in enumerate(data["modules"], start=1):
        c1 = ws.cell(r, 1, i)
        style_cell(c1, size=10, h="center", v="top")
        c2 = ws.cell(r, 2, name)
        style_cell(c2, size=10, bold=True, h="left", v="top")
        c3 = ws.cell(r, 3, desc)
        style_cell(c3, size=10, h="left", v="top")
        ws.row_dimensions[r].height = _estimate_height(
            [str(i), name, desc],
            {1: MODULE_COL_WIDTHS["A"], 2: MODULE_COL_WIDTHS["B"],
             3: MODULE_COL_WIDTHS["C"]},
        )
        r += 1

    apply_a4_print(ws, orientation="portrait", title_rows="1:4")
    ws.freeze_panes = "A5"


# === CLI ===
def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="生成报价单 Excel（参考正祥格式：深蓝表头/浅红汇总/微软雅黑）"
    )
    p.add_argument("--customer", required=True, help="客户名（用于输出目录）")
    p.add_argument("--product", required=True,
                   help="产品代号，支持逗号分隔组合（如 MI,AI）")
    p.add_argument("--mode", required=True, choices=["SAAS", "私有化"],
                   help="报价模式")
    p.add_argument("--date", default=None,
                   help="报价日期 YYYYMMDD（默认今天）")
    p.add_argument("--devkit", default=None,
                   help="可选：requirement-evaluator 二开清单路径（v1 暂未启用）")
    p.add_argument("--positions", type=int, default=3,
                   help="AI Skills 岗位数（仅 --product AI 有效，默认 3，范围 2-6）")
    args = p.parse_args(argv)
    valid_products = {"MI", "CRM", "AI"}
    product_codes = [c.strip().upper() for c in args.product.split(",") if c.strip()]
    invalid = [c for c in product_codes if c not in valid_products]
    if not product_codes or invalid:
        p.error(
            f"--product 不支持：{','.join(invalid) or args.product}；"
            f"允许值为逗号分隔组合，每个值必须在 {sorted(valid_products)} 中"
        )
    args.product_codes = product_codes
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    date_str = args.date or datetime.now().strftime("%Y%m%d")

    product_codes = args.product_codes

    products_data: list[dict[str, Any]] = []
    for code in product_codes:
        if code == "MI":
            products_data.append(MI_DATA)
        elif code == "AI":
            products_data.append(build_ai_data(args.positions))
        elif code == "CRM":
            products_data.append(CRM_DATA)

    if len(products_data) == 1:
        data = products_data[0]
    else:
        data = merge_product_data(products_data)
        data["mode_desc"] = (
            f"报价模式：{args.mode}（组合报价，"
            f"含 {len(products_data)} 个产品：{data['product_name']}）"
        )

    if args.devkit:
        if len(products_data) > 1:
            print(f"[WARN] --devkit 仅在单产品模式下生效，已忽略", file=sys.stderr)
        else:
            devkit_path = Path(args.devkit)
            if devkit_path.exists():
                dev_items = parse_devkit(devkit_path)
                if dev_items:
                    data["optional_items"] = dev_items
                    print(f"[INFO] 从需求评估加载 {len(dev_items)} 项二开")
            else:
                print(f"[WARN] --devkit 文件不存在：{devkit_path}", file=sys.stderr)

    if args.mode == "私有化":
        if not data.get("private_items"):
            print(f"[WARN] {data['product_label']} 私有化模式定价数据待确认，"
                  f"当前输出 SAAS 结构（定价可能不准）。", file=sys.stderr)
            print(f"[WARN] 请联系产品负责人确认私有化定价后补充。", file=sys.stderr)

    # 输出路径：$LANLNK_BASE/out/quotes/<客户>/报价单_<产品>_<模式>_<客户>_<日期>.xlsx
    quotes_dir = get_lanlnk_base() / "out" / "quotes"
    out_dir = quotes_dir / args.customer
    out_dir.mkdir(parents=True, exist_ok=True)
    product_str = "+".join(product_codes)
    out_file = (out_dir
                / f"报价单_{product_str}_{args.mode}_{args.customer}_{date_str}.xlsx")

    wb = Workbook()
    build_quote_sheet(wb, data, args.customer, args.mode, date_str)
    build_modules_sheet(wb, data)
    wb.save(out_file)

    print(f"[OK] 已生成报价单：{out_file}")
    print(f"     Sheet1 = {data['product_label']}报价")
    print(f"     Sheet2 = {data['product_label']}功能清单")
    return 0


if __name__ == "__main__":
    sys.exit(main())
